import ntpath
from zipfile import BadZipFile

from openpyxl import load_workbook


class Sheet:
    def __init__(self, fileName):
        self._fileName = fileName
        self._loaded = False
        self._workbook = None
        self._worksheet = None

        self._destColumn = -1
        self._codeColumn = -1
        self._rateColumn = -1
        self._dateColumn = -1
        self._startLine = -1
        self._endLine = -1

    def load(self):
        try:
            self._workbook = load_workbook(self._fileName)
            self._loaded = True
            return True
        except (BadZipFile, IOError) as e:
            print(f"An error occured while opening the file {self._fileName}: {e}")
            return False

    @property
    def filename(self):
        return ntpath.basename(self._fileName)

    @property
    def fullFilePath(self):
        return self._fileName

    @property
    def worksheet(self):
        return self._worksheet

    @property
    def workbook(self):
        return self._workbook

    @property
    def sheetnames(self):
        return self._workbook.sheetnames

    @property
    def dataFormat(self):
        return self._destColumn, self._codeColumn, self._rateColumn, self._dateColumn, self._startLine, self._endLine


    def isDataFormatDefined(self):
        return self._destColumn > 0 and self._codeColumn > 0 and self._rateColumn > 0 and self._startLine > 0 and self._worksheet is not None

    def setDataFormat(self, worksheet, destCol, codeCol, rateCol, dateCol, startLine):
        self._worksheet = self._workbook[worksheet]
        self._destColumn = destCol
        self._codeColumn = codeCol
        self._rateColumn = rateCol
        self._dateColumn = dateCol
        self._startLine = startLine
        self._endLine = self._worksheet.max_row

    @property
    def loaded(self):
        return self._loaded

    def __str__(self) -> str:
        return ntpath.basename(self._fileName)

    def printInfo(self):
        print(f"File {self.filename}:\nSheet: {self._worksheet} {self._destColumn},{self._codeColumn},{self._rateColumn}, {self._startLine}-{self._endLine}")
